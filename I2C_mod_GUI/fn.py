#file --fn.py--
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import os.path
import time

FILE = "SensoryWalk.xlsx"
#timings = []
numbers = ['zero', 'one', 'two', 'three', 'four', 'five', 'six', 'seven', 'eight', 'nine', 'ten', 'eleven', 'twelve', 'thirteen', 'fourteen', 'fifteen', 'sixteen', 'seventeen', 'eighteen', 'nineteen', 'twenty']

def reload_dictionary(user_dict):
    with open("profiles.txt") as f:
        for line in f:
            (name, lang, vol, bright) = line.split(',',4)
            user_dict[name] = {'name': name, 'lang': lang, 'vol': float(vol), 'bright': float(bright)}
    f.close()
    temp = user_dict.keys()
    sorts = sorted(temp)
    return sorts

def changeRange(value, leftMin, leftMax, rightMin, rightMax):
    # Figure out how 'wide' each range is
    leftSpan = leftMax - leftMin
    rightSpan = rightMax - rightMin

    # Convert the left range into a 0-1 range (float)
    valueScaled = float(value - leftMin) / float(leftSpan)

    # Convert the 0-1 range into a value in the right range.
    return rightMin + (valueScaled * rightSpan)

def toPlay(currentPanel, language):
    #os.system('mpg123 one.mp3 &')
    if language == 'English':
        if currentPanel == 0:
            outStr = 'mpg123 /home/pi/newGUIwI2C/Audio/English/go.mp3 && sleep 1 &&  mpg123 /home/pi/newGUIwI2C/Audio/English/zero.mp3 &'
        else:
            outStr = 'mpg123 /home/pi/newGUIwI2C/Audio/English/' + numbers[currentPanel] + '.mp3 &'
    elif language == 'Cape Verdean Creole':
        if currentPanel == 0:
            outStr = 'mpg123 /home/pi/newGUIwI2C/Audio/Cape\ Verdean\ Creole/go.mp3 && sleep 1 &&  mpg123 /home/pi/newGUIwI2C/Audio/Cape\ Verdean\ Creole/zero.mp3 &'
        else:
            outStr = 'mpg123 /home/pi/newGUIwI2C/Audio/Cape\ Verdean\ Creole/' + numbers[currentPanel] + '.mp3 &'
    elif language == 'French/Haitian Creole':
        if currentPanel == 0:
            outStr = 'mpg123 /home/pi/newGUIwI2C/Audio/French/go.mp3 && sleep 1 &&  mpg123 /home/pi/newGUIwI2C/Audio/French/zero.mp3 &'
        else:
            outStr = 'mpg123 /home/pi/newGUIwI2C/Audio/French/' + numbers[currentPanel] + '.mp3 &'
    elif language == 'Cantonese (Male)':
        if currentPanel == 0:
            outStr = 'mpg123 /home/pi/newGUIwI2C/Audio/CantoneseM/go.mp3 && sleep 1 &&  mpg123 /home/pi/newGUIwI2C/Audio/Cantonese/zero.mp3 &'
        else:
            outStr = 'mpg123 /home/pi/newGUIwI2C/Audio/CantoneseM/' + numbers[currentPanel] + '.mp3 &'
    elif language == 'Cantonese (Female)':
        if currentPanel == 0:
            outStr = 'mpg123 /home/pi/newGUIwI2C/Audio/CantoneseF/go.mp3 && sleep 1 &&  mpg123 /home/pi/newGUIwI2C/Audio/Cantonese/zero.mp3 &'
        else:
            outStr = 'mpg123 /home/pi/newGUIwI2C/Audio/CantoneseF/' + numbers[currentPanel] + '.mp3 &'
    elif language == 'Italian':
        if currentPanel == 0:
            outStr = 'mpg123 /home/pi/newGUIwI2C/Audio/Italian/go.mp3 && sleep 1 &&  mpg123 /home/pi/newGUIwI2C/Audio/Italian/zero.mp3 &'
        else:
            outStr = 'mpg123 /home/pi/newGUIwI2C/Audio/Italian/' + numbers[currentPanel] + '.mp3 &'
    elif language == 'Spanish':
        if currentPanel == 0:
            outStr = 'mpg123 /home/pi/newGUIwI2C/Audio/Spanish/go.mp3 && sleep 1 &&  mpg123 /home/pi/newGUIwI2C/Audio/Spanish/zero.mp3 &'
        else:
            outStr = 'mpg123 /home/pi/newGUIwI2C/Audio/Spanish/' + numbers[currentPanel] + '.mp3 &'

    return outStr

def excelDataSave(USR, timings):
    sheet_exists = False
    to_edit = None
    font_header = Font(name='Calibri', size=13, bold=True, italic=False, color='FF000000')
    font_data = Font(name='Calibri', size=13, bold=False, italic=False, color='FF000000')

    #if the workbook already exists
    if os.path.isfile(FILE):
        wb = load_workbook(FILE)

        #see if a sheet already exists for the user, if so, assign that sheet to to_edit
        for sh in wb:
            if sh.title == USR:
                to_edit = sh
                sheet_exists = True

        if sheet_exists:
            #print(str(to_edit.dimensions))
            temp = str(to_edit.dimensions)
            last_row = int(temp.split("X", 1)[1])
            next_row = last_row + 1

            date = to_edit.cell(row=next_row, column =1)
            date.font = font_data
            date.value = time.strftime("%m/%d/%Y")

            timer = to_edit.cell(row=next_row, column=2)
            timer.font = font_data
            timer.value = time.strftime("%I:%M:%S %p")

            feet = to_edit.cell(row=next_row, column=3)
            feet.font = font_data
            feet.value = len(timings)

            totalsec = to_edit.cell(row=next_row, column=4)
            totalsec.font = font_data
            totalsec.value = sum(timings)

            iterator = 5
            for j in timings:
                tem = to_edit.cell(row = next_row, column=iterator)
                tem.font = font_data
                tem.value = j
                iterator += 1

            current_high = int(to_edit['B2'].value)

            if len(timings) > current_high:
                to_edit['B2'].value = len(timings)

        else:
            #make a new sheet for the user
            sheet = wb.create_sheet(title=USR)
            sheet.column_dimensions['A'].width = 20
            sheet.column_dimensions['B'].width = 20

             #set up headers and things
            sheet['A1'] = 'Name'
            sheet['A1'].font = font_header
            sheet['B1'] = USR
            sheet['B1'].font = font_data
            sheet['A2'] = 'Personal Best in Feet'
            sheet['A2'].font = font_header
            #this is a temporary number, on first load, this doesn't have a number, so the number of feet just achieved is put here
            sheet['B2'] = len(timings)
            sheet['B2'].font = font_data

            sheet['A4'] = 'Date'
            sheet['A4'].font = font_header
            sheet['B4'] = 'Time of Day'
            sheet['B4'].font = font_header
            sheet['C4'] = 'Total Feet'
            sheet['C4'].font = font_header
            sheet['D4'] = 'Total Time'
            sheet['D4'].font = font_header
            iter = 0

            for col in range(5,25):
                temp = sheet.cell(row = 4, column = col)
                temp.value = str(iter) + ' to ' + str(iter+1)
                temp.font = font_header
                iter += 1

            #actually print the data here
            date = sheet.cell(row=5, column =1)
            date.font = font_data
            date.value = time.strftime("%m/%d/%Y")

            timer = sheet.cell(row=5, column=2)
            timer.font = font_data
            timer.value = time.strftime("%I:%M:%S %p")

            feet = sheet.cell(row=5, column=3)
            feet.font = font_data
            feet.value = len(timings)

            totalsec = sheet.cell(row=5, column=4)
            totalsec.font = font_data
            totalsec.value = sum(timings)

            iterator = 5
            for j in timings:
                tem = sheet.cell(row = 5, column=iterator)
                tem.font = font_data
                tem.value = j
                iterator += 1

    #if the workbook doesn't exist
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.title = USR
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20

        #set up headers and things
        sheet['A1'] = 'Name'
        sheet['A1'].font = font_header
        sheet['B1'] = USR
        sheet['B1'].font = font_data
        sheet['A2'] = 'Personal Best in Feet'
        sheet['A2'].font = font_header
        #this is a temporary number, on first load, this doesn't have a number, so the number of feet just achieved is put here
        sheet['B2'] = len(timings)
        sheet['B2'].font = font_data

        sheet['A4'] = 'Date'
        sheet['A4'].font = font_header
        sheet['B4'] = 'Time of Day'
        sheet['B4'].font = font_header
        sheet['C4'] = 'Total Feet'
        sheet['C4'].font = font_header
        sheet['D4'] = 'Total Time'
        sheet['D4'].font = font_header
        iter = 0

        for col in range(5,25):
            temp = sheet.cell(row = 4, column = col)
            temp.value = str(iter) + ' to ' + str(iter+1)
            temp.font = font_header
            iter += 1

        #actually print the data here
        date = sheet.cell(row=5, column =1)
        date.font = font_data
        date.value = time.strftime("%m/%d/%Y")

        timer = sheet.cell(row=5, column=2)
        timer.font = font_data
        timer.value = time.strftime("%I:%M:%S %p")

        feet = sheet.cell(row=5, column=3)
        feet.font = font_data
        feet.value = len(timings)

        totalsec = sheet.cell(row=5, column=4)
        totalsec.font = font_data
        totalsec.value = sum(timings)

        iterator = 5
        for j in timings:
            tem = sheet.cell(row = 5, column=iterator)
            tem.font = font_data
            tem.value = j
            iterator += 1


    wb.save(FILE)

